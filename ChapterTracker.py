"""
Chapter Tracker
Author: Connor Maclachlan
"""

import pandas as pd
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
import tkinter as tk
from tkinter import ttk
from tkinter import filedialog
from ttkthemes import ThemedTk
from tkinter import messagebox


# Chapter Tracker class 
class ChapterTracker:

    def __init__(self) -> None:

        self.root = ThemedTk(theme="plastik")
        self.root.title('ChapterTracker')

        self.HomePage = ttk.Frame(self.root)

        self.checkInCSV = None
        self.excusedCSV = None
        self.members = []

        self.attendanceFP = None
        self.attendanceWB = None
        self.activeSheet = None
        self.sheetType = tk.IntVar()

        self.rows = None
        self.columns = None
        self.currentColumn = None

        self.HomePage.pack(fill="both", expand=1)

        # Home page headers
        self.headlabel1 = ttk.Label(self.HomePage, text="Chapter Tracker", font=('Arial', 20))
        self.headlabel1.pack(padx=10, pady=10)

        self.headlabel2 = ttk.Label(self.HomePage, text="To Begin, Please Select A Spreadsheet:", font=('Arial', 18))
        self.headlabel2.pack(padx=10, pady=10)

        # Button to select the current attendance sheet
        self.spreadsheetSelect = ttk.Button(self.HomePage, text="Select Spreadsheet", command=lambda: self.openFile("Spreadsheet"))
        self.spreadsheetSelect.pack(padx=10, pady=10)

        # Continue label and button
        self.contlabel = ttk.Label(self.HomePage, text="Click Continue To Proceed", font=('Arial', 18))
        self.contlabel.pack(padx=10, pady=10)

        self.continueBtn = ttk.Button(self.HomePage, text="Continue", command=self.openEditPage)
        self.continueBtn.pack(padx=10, pady=10)

        self.root.mainloop()

    def openEditPage(self):
        """ 
            Trasitions from home page to editing page and 
            initializes editing page elements
        """
        self.EditPage = ttk.Frame(self.root)
        
        if self.attendanceWB is not None:
            self.EditPage.pack(fill="both", expand=1)
            self.HomePage.pack_forget()

            widgets = ttk.LabelFrame(self.EditPage, text="Enter Data")
            widgets.grid(row=0, column=0, padx=10, pady=10)

            self.eventName = ttk.Entry(widgets)
            self.eventName.insert(0, "Event Name")
            self.eventName.bind("<FocusIn>", lambda e: self.eventName.delete(0, 'end'))
            self.eventName.grid(row=0, column=0, sticky='ew', padx=10, pady=5)

            self.eventDate = ttk.Entry(widgets)
            self.eventDate.insert(0, "DD/MM/YYYY")
            self.eventDate.bind("<FocusIn>", lambda e: self.eventDate.delete(0, 'end'))
            self.eventDate.grid(row=1, column=0, sticky='ew', padx=10, pady=5)

            chapterSheetRadioBtn = ttk.Radiobutton(widgets, text='Chapter', variable=self.sheetType, value=0)
            chapterSheetRadioBtn.grid(row=2, column=0, sticky='ew', padx=10, pady=5)

            eventSheetRadioBtn = ttk.Radiobutton(widgets, text='Event', variable=self.sheetType, value=1)
            eventSheetRadioBtn.grid(row=3, column=0, sticky='ew', padx=10, pady=5)

            checkInBtn = ttk.Button(widgets, text="Upload Check-In CSV", command=lambda: self.openFile("CheckIn"))
            checkInBtn.grid(row=4, column=0, sticky='ew', padx=10, pady=5)

            abscenceBtn = ttk.Button(widgets, text="Upload Abscence CSV", command=lambda: self.openFile("Absence"))
            abscenceBtn.grid(row=5, column=0, sticky='ew', padx=10, pady=5)

            seperator = ttk.Separator(widgets)
            seperator.grid(row=6, column=0, padx=10, pady=10, sticky="ew")

            trackBtn = ttk.Button(widgets, text="Track Attendance", command=self.recordAttendance)
            trackBtn.grid(row=7, column=0, sticky='ew', padx=10, pady=5)

            unknownNamesFrame = ttk.LabelFrame(self.EditPage, text="Unknown Names")
            unknownNamesFrame.grid(row=0, column=1, padx=10, pady=10)

            self.unknownNamesText = tk.Text(unknownNamesFrame, width=20, height=18)
            self.unknownNamesText.pack(padx=10, pady=5)

            backBtn = ttk.Button(self.EditPage, text="Back", command=self.openHomePage)
            backBtn.grid(row=1, column=1, sticky='ew', padx=10, pady=5)


        else:
            messagebox.showerror("Error", "Please Select An Attendance Spreadsheet")

    def openHomePage(self):
        """ Returns to the home page """
        self.EditPage.destroy()
        self.HomePage.pack(fill="both", expand=1)

    def openFile(self, fileType):
        """ Opens selected files and initializes file pointers """

        if fileType == "Spreadsheet":
            self.attendanceFP = filedialog.askopenfilename(initialdir='/Desktop', title='Select A File', filetypes=[('Excel Files', '*.xlsx')])
            self.attendanceWB = load_workbook(self.attendanceFP)
            
        elif fileType == "CheckIn":
            self.checkInCSV = pd.read_csv(filedialog.askopenfilename(initialdir='/Desktop', title='Select A File', filetypes=[('CSV Files', '*.csv')]))

        else:
            self.excusedCSV = pd.read_csv(filedialog.askopenfilename(initialdir='/Desktop', title='Select A File', filetypes=[('CSV Files', '*.csv')]))

    def processName(self, name):
        """ Processes names from file to reduce inconsitencies and errors """

        if " " in name:
            processedName = name.lower().strip().split()
            firstInitial = processedName[0][0]
            lastName = processedName[1]
            processedName = firstInitial + " " + lastName
            return processedName
        
        else:
            return name
    
    def selectSheet(self):
        """ Sets desired workbook sheet to the active sheet """

        sheet = None
        if self.sheetType.get() == 0:
            sheet = 'Chapters'
        elif self.sheetType.get() == 1:
            sheet = 'Events'

        self.activeSheet = self.attendanceWB[sheet]
        self.rows = self.activeSheet.max_row
        self.columns = self.activeSheet.max_column

    def addEvent(self):
        """ Creates new event in active sheet """

        self.currentColumn = chr(65 + self.columns)
        self.activeSheet[self.currentColumn + '2'] = self.eventName.get() + ' ' + self.eventDate.get()

    def reset(self):
        """ Resets class attributes """

        self.checkInCSV = None
        self.excusedCSV = None
        self.members = []
        self.rows = None
        self.columns = None
        self.currentColumn = None

    def recordAttendance(self):
        """ 
            Processes and records attendance data from various CSV files 
            into the desired spreadsheet
        """
        if self.checkInCSV is None:
            messagebox.showerror("Error", "Please Select A Check-In CSV File")
        
        elif self.excusedCSV is None:
            messagebox.showerror("Error", "Please Select An Excused Absence CSV File")
        
        else:
            self.unknownNamesText.config(state=tk.NORMAL)
            self.unknownNamesText.delete('1.0', tk.END)
            
            self.selectSheet()
            self.addEvent()

            for cell in self.activeSheet['A']:
                self.members.append(cell.value)

            memberList = self.members[2:]
            checkInList = self.checkInCSV['First And Last Name'].to_list()
            excusedList = self.excusedCSV['Full Name'].to_list()

            processedMembers = []
            processedCheckIn = []
            processedExcused = []

            for i in range(len(memberList)):
                if memberList[i] is not None:
                    processedMembers.append(self.processName(memberList[i]))

            for i in range(len(checkInList)):
                if checkInList[i] is not None:
                    processedCheckIn.append(self.processName(checkInList[i]))

            for i in range(len(excusedList)):
                if excusedList[i] is not None:
                    processedExcused.append(self.processName(excusedList[i]))

            missing = []
            absent = []

            for i in range(len(processedMembers)):

                # Mark as absent
                if (processedMembers[i] not in processedCheckIn) and (processedMembers[i] not in processedExcused):
                    absent.append(memberList[i])
                    self.activeSheet[self.currentColumn + str(i+3)] = 'A'
                    print(memberList[i])

                # Mark as excused
                elif (processedMembers[i] not in processedCheckIn) and (processedMembers[i]  in processedExcused):
                    self.activeSheet[self.currentColumn + str(i+3)] = 'E'

                # Mark as present
                elif (processedMembers[i] in processedCheckIn):
                    self.activeSheet[self.currentColumn + str(i+3)] = 'P'

            for i in range(len(processedCheckIn)):
                if processedCheckIn[i] not in processedMembers:
                    missing.append(checkInList[i])
                    self.unknownNamesText.insert(tk.END, checkInList[i] +'\n')

            self.attendanceWB.save(self.attendanceFP)
            messagebox.showinfo("Success!", "The Data Has Been Saved Successfully")
            self.reset()

ChapterTracker()
